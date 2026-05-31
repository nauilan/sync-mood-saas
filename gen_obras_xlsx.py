"""
Parse mock-obras.ts and generate CWR_TSM_760_OBRAS.xlsx
"""
import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TS_FILE = r"C:\Users\Usuário\Desktop\sync-mood-saas\apps\web\lib\mock-obras.ts"
OUT_FILE = r"C:\Users\Usuário\Desktop\CWR_TSM_760_OBRAS.xlsx"

print("Reading TS file...")
with open(TS_FILE, encoding="utf-8") as f:
    content = f.read()

# ------------------------------------------------------------------ #
# Split into individual obra blocks by looking for top-level { ... } #
# ------------------------------------------------------------------ #

marker = content.find("export const MOCK_OBRAS: Obra[] = [")
assert marker != -1, "Array start not found"
# Find '= [' after the marker to get the outer array bracket, not 'Obra[]'
eq_bracket = content.index("= [", marker)
array_start = eq_bracket + 2 + 1  # skip past '= ['

obras_raw = []
depth = 0
current_start = None
i = array_start

while i < len(content):
    ch = content[i]
    if ch == '{':
        if depth == 0:
            current_start = i
        depth += 1
    elif ch == '}':
        depth -= 1
        if depth == 0 and current_start is not None:
            obras_raw.append(content[current_start:i+1])
            current_start = None
    elif ch == ']' and depth == 0:
        break
    i += 1

print(f"  Found {len(obras_raw)} raw obra blocks")

# ------------------------------------------------------------------ #
# Helper: extract a simple scalar field                               #
# ------------------------------------------------------------------ #

def get_field(block, key):
    m = re.search(rf'(?<!\w){re.escape(key)}\s*:\s*"([^"]*)"', block)
    if m:
        return m.group(1)
    m = re.search(rf'(?<!\w){re.escape(key)}\s*:\s*([^,\n\]}}]+)', block)
    if m:
        return m.group(1).strip().rstrip(',').strip()
    return ""

# ------------------------------------------------------------------ #
# Extract sub-array block (first [ ... ] for a given key)            #
# ------------------------------------------------------------------ #

def extract_array(text, key):
    match = re.search(rf'\b{re.escape(key)}\s*:\s*\[', text)
    if not match:
        return ""
    start = match.end() - 1
    depth = 0
    for j in range(start, len(text)):
        c = text[j]
        if c == '[':
            depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0:
                return text[start:j+1]
    return text[start:]

def extract_objects(array_str):
    """Return list of top-level {...} strings from an array string."""
    objects = []
    inner = array_str[1:-1] if array_str.startswith('[') else array_str
    depth = 0
    start = None
    for k, c in enumerate(inner):
        if c == '{':
            if depth == 0:
                start = k
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0 and start is not None:
                objects.append(inner[start:k+1])
                start = None
    return objects

# ------------------------------------------------------------------ #
# Parse each obra block                                               #
# ------------------------------------------------------------------ #

def parse_obra(block):
    obra = {}
    obra["codigo"]   = get_field(block, "codigo")
    obra["titulo"]   = get_field(block, "titulo")
    obra["iswc"]     = get_field(block, "iswc")
    obra["duracao"]  = get_field(block, "duracao")
    obra["genero"]   = get_field(block, "genero")
    obra["ano"]      = get_field(block, "ano")
    obra["idioma"]   = get_field(block, "idioma")
    obra["status"]   = get_field(block, "status")
    obra["observacoes"] = get_field(block, "observacoes")
    obra["id"]       = get_field(block, "id")
    obra["_percentual_controlado"] = get_field(block, "_percentual_controlado")

    autores_list  = []
    editoras_list = []

    links_array = extract_array(block, "_links")
    link_objects = extract_objects(links_array)

    for lobj in link_objects:
        tit_array = extract_array(lobj, "titulares")
        tit_objects = extract_objects(tit_array)

        for tobj in tit_objects:
            nome  = get_field(tobj, "nome")
            papel = get_field(tobj, "papel").lower()
            pct   = get_field(tobj, "percentual")
            ctrl  = get_field(tobj, "controlado")
            try:
                pct_str = f"{float(pct):.2f}%"
            except Exception:
                pct_str = f"{pct}%"

            is_editor = any(x in papel for x in [
                "editora", "editor", "administradora", "sub_editora",
                "subeditor", "administrador", "sub_editor"
            ])
            is_autor = any(x in papel for x in [
                "autor", "compositor", "lyricist", "writer", "co_autor",
                "co_compositor"
            ])

            entry = f"{nome} ({pct_str})"
            if is_editor:
                editoras_list.append(f"{entry} [controlado={ctrl}]")
            elif is_autor:
                autores_list.append(entry)
            else:
                autores_list.append(f"{entry} [{papel}]")

    obra["autores"]  = "; ".join(autores_list)
    obra["editoras"] = "; ".join(editoras_list)
    return obra

print("Parsing obra blocks...")
obras = [parse_obra(b) for b in obras_raw]
print(f"  Parsed {len(obras)} obras")

iswc_count = sum(1 for o in obras if o["iswc"])
print(f"  ISWC filled: {iswc_count}/{len(obras)}")

# ------------------------------------------------------------------ #
# Write Excel                                                         #
# ------------------------------------------------------------------ #
print("Writing Excel...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Obras CWR"

HEADERS = [
    "Codigo", "Titulo", "ISWC", "Duracao", "Idioma",
    "Autores", "Editoras", "Genero", "Ano", "Status",
    "Observacoes", "% Controlado", "ID"
]

FIELD_MAP = [
    "codigo", "titulo", "iswc", "duracao", "idioma",
    "autores", "editoras", "genero", "ano", "status",
    "observacoes", "_percentual_controlado", "id"
]

header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
header_fill  = PatternFill("solid", fgColor="1F4E79")
header_align = Alignment(horizontal="center", vertical="center")

for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align

data_align = Alignment(vertical="top", wrap_text=False)
wrap_align  = Alignment(vertical="top", wrap_text=True)

for row_idx, obra in enumerate(obras, start=2):
    for col_idx, field in enumerate(FIELD_MAP, start=1):
        val = obra.get(field, "") or None
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if field in ("autores", "editoras", "observacoes"):
            cell.alignment = wrap_align
        else:
            cell.alignment = data_align

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

COL_WIDTHS = {
    "A": 10,  "B": 35,  "C": 18,  "D": 10,  "E": 8,
    "F": 55,  "G": 55,  "H": 15,  "I": 8,   "J": 10,
    "K": 40,  "L": 14,  "M": 20,
}
for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 22

wb.save(OUT_FILE)
size = os.path.getsize(OUT_FILE)
print(f"\nDone!")
print(f"  File : {OUT_FILE}")
print(f"  Size : {size:,} bytes  ({size/1024:.1f} KB)")
print(f"  Rows : {len(obras)} obras + 1 header = {len(obras)+1} total")
print(f"  ISWC : {iswc_count}/{len(obras)} preenchidos")
