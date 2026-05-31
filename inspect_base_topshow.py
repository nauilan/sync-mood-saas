import pandas as pd
import openpyxl

# Inspecionar estrutura
wb = openpyxl.load_workbook(r"C:\Users\Usuário\Downloads\base top show (1).xlsx", read_only=True, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"[{name}] rows={ws.max_row} cols={ws.max_column}")
wb.close()

df = pd.read_excel(r"C:\Users\Usuário\Downloads\base top show (1).xlsx", sheet_name=None, nrows=0)
for name, d in df.items():
    print(f"[{name}] colunas: {list(d.columns)}")

print("\n--- Primeiras linhas ---")
df_full = pd.read_excel(r"C:\Users\Usuário\Downloads\base top show (1).xlsx", dtype=str)
print(df_full.head(10).to_string())
